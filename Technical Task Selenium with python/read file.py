import openpyxl

file="C:\Book1.xlsx"

wb = openpyxl.load_workbook(file)

sheet = wb.active

row = sheet.max_row
col = sheet.max_column

for r in range(1,row+1):
    for c in range(1,col+1):
        print(sheet.cell(r,c).value,end='')
    print()